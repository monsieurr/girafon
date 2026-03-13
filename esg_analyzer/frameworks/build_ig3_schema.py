from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


SHEETS = {
    "ESRS 2": {"category": "General", "pillar": "General Disclosures"},
    "ESRS 2 MDR": {"category": "General", "pillar": "Minimum Disclosure Requirements"},
    "ESRS E1": {"category": "Environment", "pillar": "Climate Change"},
    "ESRS E2": {"category": "Environment", "pillar": "Pollution"},
    "ESRS E3": {"category": "Environment", "pillar": "Water and Marine Resources"},
    "ESRS E4": {"category": "Environment", "pillar": "Biodiversity and Ecosystems"},
    "ESRS E5": {"category": "Environment", "pillar": "Resource Use and Circular Economy"},
    "ESRS S1": {"category": "Social", "pillar": "Own Workforce"},
    "ESRS S2": {"category": "Social", "pillar": "Workers in the Value Chain"},
    "ESRS S3": {"category": "Social", "pillar": "Affected Communities"},
    "ESRS S4": {"category": "Social", "pillar": "Consumers and End-users"},
    "ESRS G1": {"category": "Governance", "pillar": "Business Conduct"},
}

STOPWORDS = {
    "the", "and", "or", "of", "to", "in", "for", "with", "by", "on", "from", "at",
    "as", "is", "are", "be", "been", "being", "this", "that", "these", "those",
    "including", "including", "related", "relation", "about", "such", "any", "all",
    "own", "its", "their", "undertaking", "entity", "companies", "company", "group",
    "disclosure", "disclosures", "information", "statement", "statements",
}

DATA_TYPE_UNITS = {
    "percent": ["%"],
    "percentage": ["%"],
    "monetary": ["EUR"],
    "number": [],
    "integer": [],
    "decimal": [],
    "narrative": [],
    "text": [],
}


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


def _find_header_indices(header: List[Any]) -> Dict[str, int]:
    normed = [_norm(h) for h in header]
    idx = {}

    def find(needle: str) -> Optional[int]:
        for i, h in enumerate(normed):
            if needle in h:
                return i
        return None

    mapping = {
        "id": "id",
        "esrs": "esrs",
        "dr": "dr",
        "paragraph": "paragraph",
        "related ar": "related ar",
        "name": "name",
        "data type": "data type",
        "conditional": "conditional or alternative",
        "may": "may",
        "appendix b": "appendix b",
        "appendix c <750": "appendix c - esrs 1 dps subject to phasing-in provisions applicable to undertaking with less than 750 employees",
        "appendix c all": "appendix c - esrs 1 dps subject to phasing-in provisions applicable to all undertakings",
    }

    for key, needle in mapping.items():
        idx_val = find(needle)
        if idx_val is not None:
            idx[key] = idx_val

    required = ["id", "esrs", "dr", "name", "data type"]
    missing = [r for r in required if r not in idx]
    if missing:
        raise ValueError(f"Missing required columns in IG3 sheet: {missing}")
    return idx


def _token_keywords(name: str) -> List[str]:
    tokens = re.findall(r"[A-Za-z0-9]+", name.lower())
    cleaned = []
    for t in tokens:
        if t in STOPWORDS:
            continue
        if len(t) < 3 and t not in {"ghg", "co2", "co2e"}:
            continue
        cleaned.append(t)
    return cleaned[:6]


def _build_keywords(name: str, dr: str) -> List[str]:
    keywords: List[str] = []
    phrase = str(name or "").strip()
    if phrase:
        keywords.append(phrase)
    if dr:
        keywords.append(dr.strip())
    keywords.extend(_token_keywords(phrase))

    seen = set()
    uniq = []
    for kw in keywords:
        if not kw:
            continue
        if kw in seen:
            continue
        seen.add(kw)
        uniq.append(kw)
    return uniq


def _data_type_units(data_type: str) -> List[str]:
    if not data_type:
        return []
    return DATA_TYPE_UNITS.get(_norm(data_type), [])


def _is_quantitative(data_type: str) -> bool:
    dt = _norm(data_type)
    return dt not in {"narrative", "text"}


def _scoring_config_fallback() -> Dict[str, Any]:
    return {
        "bands": {
            "excellent": {"min": 80, "max": 100, "label": "Excellent", "color": "#22c55e"},
            "good": {"min": 60, "max": 79, "label": "Good", "color": "#84cc16"},
            "needs_improvement": {"min": 40, "max": 59, "label": "Needs Improvement", "color": "#f59e0b"},
            "weak": {"min": 0, "max": 39, "label": "Weak / High Risk", "color": "#ef4444"},
        },
        "greenwashing_flags": [],
    }


def build_schema(ig3_path: Path, base_schema_path: Optional[Path] = None) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(ig3_path, read_only=True, data_only=True)

    scoring_config = None
    if base_schema_path and base_schema_path.exists():
        try:
            scoring_config = json.loads(base_schema_path.read_text(encoding="utf-8")).get("_scoring_config")
        except Exception:
            scoring_config = None
    if not scoring_config:
        scoring_config = _scoring_config_fallback()

    schema: Dict[str, Any] = {
        "_meta": {
            "version": "IG3-1.0",
            "source": "EFRAG IG3 List of ESRS Data Points (Excel)",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "note": "IG3 datapoints are mostly subject to materiality; use results as coverage signals, not full compliance.",
        },
        "_scoring_config": scoring_config,
    }

    for sheet_name, meta in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet '{sheet_name}' in IG3 workbook.")
        ws = wb[sheet_name]

        header = [c for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
        idx = _find_header_indices(header)

        for row in ws.iter_rows(min_row=3, values_only=True):
            dp_id = row[idx["id"]]
            if not dp_id:
                continue

            dp_id = str(dp_id).strip()
            # Skip non-datapoint header lines that appear in the sheets
            if "_" not in dp_id:
                continue
            if not re.search(r"\d", dp_id):
                continue
            esrs = str(row[idx["esrs"]] or "").strip()
            dr = str(row[idx["dr"]] or "").strip()
            name = str(row[idx["name"]] or "").strip()
            data_type = str(row[idx["data type"]] or "").strip()

            conditional = str(row[idx.get("conditional")] or "").strip() if idx.get("conditional") is not None else ""
            voluntary = str(row[idx.get("may")] or "").strip() if idx.get("may") is not None else ""
            appendix_b = str(row[idx.get("appendix b")] or "").strip() if idx.get("appendix b") is not None else ""
            appendix_c_750 = str(row[idx.get("appendix c <750")] or "").strip() if idx.get("appendix c <750") is not None else ""
            appendix_c_all = str(row[idx.get("appendix c all")] or "").strip() if idx.get("appendix c all") is not None else ""

            is_voluntary = bool(voluntary)
            is_conditional = bool(conditional)
            is_mandatory = (sheet_name in {"ESRS 2", "ESRS 2 MDR"}) and not is_voluntary

            requirement_type = "mandatory" if is_mandatory else "voluntary" if is_voluntary else "conditional" if is_conditional else "materiality"

            weight = 0 if is_voluntary else 1

            schema[dp_id] = {
                "framework": "ESRS",
                "section": dr,
                "name": name,
                "category": meta["category"],
                "pillar": meta["pillar"],
                "description": name,
                "keywords": _build_keywords(name, dr),
                "expected_units": _data_type_units(data_type),
                "expected_data_points": [],
                "quality_checks": [],
                "is_quantitative": _is_quantitative(data_type),
                "original_mandatory": is_mandatory,
                "omnibus_mandatory": is_mandatory,
                "omnibus_notes": "",
                "weight_original": weight,
                "weight_omnibus": weight,
                "ig3": {
                    "datapoint_id": dp_id,
                    "esrs": esrs,
                    "data_type": data_type,
                    "conditional_or_alternative": conditional,
                    "voluntary": is_voluntary,
                    "appendix_b": appendix_b,
                    "appendix_c_lt_750": appendix_c_750,
                    "appendix_c_all": appendix_c_all,
                    "requirement_type": requirement_type,
                },
            }

    return schema


def main() -> None:
    parser = argparse.ArgumentParser(description="Build IG3 ESRS schema JSON from EFRAG workbook.")
    parser.add_argument("--ig3", required=True, help="Path to EFRAG IG3 Excel workbook")
    parser.add_argument("--out", required=True, help="Output JSON path")
    parser.add_argument("--base-schema", default="", help="Optional base schema to copy _scoring_config from")
    args = parser.parse_args()

    ig3_path = Path(args.ig3)
    if not ig3_path.exists():
        raise FileNotFoundError(f"IG3 workbook not found: {ig3_path}")

    base_schema_path = Path(args.base_schema) if args.base_schema else None

    schema = build_schema(ig3_path, base_schema_path)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(schema, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"OK  IG3 schema written: {out_path}")


if __name__ == "__main__":
    main()
