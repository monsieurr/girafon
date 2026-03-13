from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List

import openpyxl


EXCLUDE_LABEL_SUFFIX = re.compile(r"\[(axis|member|table|line items|abstract)\]$", re.I)

# Mapping rules for the disclosures currently in esrs_schema.json (20 keys).
RULES: Dict[str, Dict[str, Any]] = {
    "E1-6_Scope1": {
        "role_contains": ["E1-6"],
        "include": ["scope 1"],
        "exclude": ["scope 2", "scope 3"],
    },
    "E1-6_Scope2": {
        "role_contains": ["E1-6"],
        "include": ["scope 2"],
        "exclude": ["scope 1", "scope 3"],
    },
    "E1-6_Scope3": {
        "role_contains": ["E1-6"],
        "include": ["scope 3"],
        "exclude": ["scope 1", "scope 2"],
    },
    "E1-6_Intensity": {
        "role_contains": ["E1-6.1", "E1-6"],
        "include": ["intensity"],
        "exclude": [],
    },
    "E1-4_Targets": {"role_contains": ["E1-4"], "include": [], "exclude": []},
    "E1-5_Energy": {"role_contains": ["E1-5"], "include": [], "exclude": []},
    "E1-1_TransitionPlan": {"role_contains": ["E1-1"], "include": [], "exclude": []},
    "E2-4_Pollution": {"role_contains": ["E2-4"], "include": [], "exclude": []},
    "E3-4_Water": {"role_contains": ["E3-4"], "include": [], "exclude": []},
    "E4-1_Biodiversity": {"role_contains": ["E4-1"], "include": [], "exclude": []},
    "E5-5_Waste": {"role_contains": ["E5-5"], "include": [], "exclude": []},
    "S1-9_Diversity": {"role_contains": ["S1-9"], "include": [], "exclude": []},
    "S1-14_HealthSafety": {"role_contains": ["S1-14"], "include": [], "exclude": []},
    "S1-13_Training": {"role_contains": ["S1-13"], "include": [], "exclude": []},
    "S1-8_CollectiveBargaining": {"role_contains": ["S1-8"], "include": [], "exclude": []},
    "S2-1_SupplyChain": {"role_contains": ["S2-1"], "include": [], "exclude": []},
    "G1-1_Governance": {"role_contains": ["G1-1"], "include": [], "exclude": []},
    "G1-3_AntiCorruption": {"role_contains": ["G1-3"], "include": [], "exclude": []},
    "G1-4_CorruptionIncidents": {"role_contains": ["G1-4"], "include": [], "exclude": []},
    "G1-5_PoliticalInfluence": {"role_contains": ["G1-5"], "include": [], "exclude": []},
}


def _is_data_row(label: str, technical: str) -> bool:
    if not label or not technical:
        return False
    if EXCLUDE_LABEL_SUFFIX.search(label):
        return False
    if technical.endswith(("Axis", "Member", "Table", "LineItems", "Abstract")):
        return False
    return True


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def _load_rows(annex_path: Path, sheet: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(annex_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: i for i, name in enumerate(header)}
    required = [
        "Role",
        "Label en",
        "Technical Name",
        "Abstract",
        "Type name short",
        "Period type",
        "Balance",
        "Substitution Group",
        "References",
    ]
    for name in required:
        if name not in col:
            raise ValueError(f"Missing column '{name}' in sheet '{sheet}'.")

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        role = r[col["Role"]]
        label = r[col["Label en"]]
        technical = r[col["Technical Name"]]
        if not _is_data_row(label, technical):
            continue
        rows.append(
            {
                "role": role or "",
                "label": label or "",
                "label_lc": _normalize(label or ""),
                "technical_name": technical or "",
                "abstract": r[col["Abstract"]],
                "type_name_short": r[col["Type name short"]],
                "period_type": r[col["Period type"]],
                "balance": r[col["Balance"]],
                "substitution_group": r[col["Substitution Group"]],
                "references": r[col["References"]] or "",
            }
        )
    return rows


def _filter_rows(rows: Iterable[Dict[str, Any]], rule: Dict[str, Any]) -> List[Dict[str, Any]]:
    role_contains = rule.get("role_contains", [])
    include = [i.lower() for i in rule.get("include", [])]
    exclude = [e.lower() for e in rule.get("exclude", [])]

    filtered = [
        r for r in rows
        if any(rc in r["role"] for rc in role_contains)
    ]
    if include:
        filtered = [
            r for r in filtered
            if any(tok in r["label_lc"] for tok in include)
        ]
    if exclude:
        filtered = [
            r for r in filtered
            if not any(tok in r["label_lc"] for tok in exclude)
        ]

    # Dedupe by technical name
    deduped = {}
    for r in filtered:
        deduped[r["technical_name"]] = r
    return list(deduped.values())


def build_mapping(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    by_disclosure: Dict[str, Any] = {}
    for key, rule in RULES.items():
        elements = _filter_rows(rows, rule)
        by_disclosure[key] = {
            "elements": sorted(elements, key=lambda x: x["technical_name"]),
            "match_rule": rule,
        }

    by_taxonomy: Dict[str, Any] = {}
    for key, payload in by_disclosure.items():
        for el in payload["elements"]:
            by_taxonomy[el["technical_name"]] = {
                "disclosure_key": key,
                "role": el["role"],
                "label": el["label"],
            }

    return {
        "_meta": {
            "source": "EFRAG ESRS Set 1 XBRL Taxonomy Annex 1 (Excel)",
            "sheet": "PresentationLinkbase",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "rules_version": "1.0",
            "notes": "Built from PresentationLinkbase; axes/members/tables/abstracts excluded.",
        },
        "by_disclosure": by_disclosure,
        "by_taxonomy_element": by_taxonomy,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build ESRS taxonomy mapping JSON from EFRAG Annex 1.")
    parser.add_argument("--annex", required=True, help="Path to Annex 1 ESRS Set 1 XBRL Taxonomy Excel file")
    parser.add_argument("--out", required=True, help="Output JSON path")
    parser.add_argument("--sheet", default="PresentationLinkbase", help="Sheet name to parse")
    args = parser.parse_args()

    annex_path = Path(args.annex)
    if not annex_path.exists():
        raise FileNotFoundError(f"Annex file not found: {annex_path}")

    rows = _load_rows(annex_path, args.sheet)
    mapping = build_mapping(rows)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(mapping, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"OK  Mapping written: {out_path}")


if __name__ == "__main__":
    main()
